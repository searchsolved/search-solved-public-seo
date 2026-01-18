####################################################################################
#                                                                                  #
#  Content Decay Analyzer                                                          #
#                                                                                  #
#  Analyze GSC data to find pages losing traffic over time.                        #
#                                                                                  #
####################################################################################
# Author   : Lee Foot                                                              #
# Website  : https://www.leefoot.com                                                   #
# Contact  : https://www.leefoot.com/contact                                           #
# Email    : hello@leefoot.com                                                     #
# LinkedIn : https://www.linkedin.com/in/lee-foot/                                 #
# Bluesky  : https://bsky.app/profile/leefootseo.bsky.social                       #
####################################################################################

"""
Content Decay Analyzer

Analyzes Google Search Console data to identify pages experiencing traffic decline.
Compares peak month performance vs latest month to calculate clicks lost.
Generates Excel reports with color-coded formatting.

Features:
- Upload GSC data CSV with date, page, query, clicks columns
- Configurable analysis period (months)
- Identifies peak performance month for each page
- Calculates clicks lost from peak to current
- Exports formatted Excel with conditional formatting
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Content Decay Analyzer", page_icon="📉", layout="wide")

st.title("Content Decay Analyzer")
st.markdown("*Created by* [![Website](https://img.shields.io/badge/-leefoot.com-2A9D8F?logoColor=white)](https://www.leefoot.com) · [![Hire Me](https://img.shields.io/badge/-Hire%20Me-FF6B6B?logoColor=white)](https://www.leefoot.com/contact) · [![LinkedIn](https://img.shields.io/badge/-LinkedIn-0A66C2?logo=linkedin&logoColor=white)](https://www.linkedin.com/in/lee-foot/) · [![Bluesky](https://img.shields.io/badge/-Bluesky-0285FF?logo=bluesky&logoColor=white)](https://bsky.app/profile/leefootseo.bsky.social) · [![More Tools](https://img.shields.io/badge/-More%20Tools-8B5CF6?logoColor=white)](https://leefoot.com/tools) · [![GitHub](https://img.shields.io/badge/-GitHub-181717?logo=github&logoColor=white)](https://github.com/searchsolved/search-solved-public-seo)")

with st.expander("How to use this tool"):
    st.markdown("""
    **What this tool does:**
    - Identifies pages that have lost traffic compared to their peak performance
    - Compares monthly click data to find content decay patterns
    - Highlights which months had the highest clicks for each page

    **Data requirements:**
    Your GSC export should have these columns:
    - `date` - Date of the data (e.g., 2024-01-15)
    - `page` - The URL of the page
    - `query` - The search query (optional, used for aggregation)
    - `clicks` - Number of clicks

    **How to get this data:**
    1. Export from Google Search Console (Performance > Export)
    2. Use the Search Console API with date, page, query dimensions
    3. Use GSC connector tools

    **Clicks Lost Explained:**
    The 'Clicks Lost' metric shows the difference between a page's peak month
    clicks and its latest month clicks. Negative values indicate content decay.
    """)

# Sidebar settings
st.sidebar.header("Analysis Settings")

months_to_analyze = st.sidebar.slider(
    "Months to analyze",
    min_value=3,
    max_value=24,
    value=12,
    help="Number of most recent complete months to include in analysis"
)

min_peak_clicks = st.sidebar.number_input(
    "Minimum peak clicks",
    min_value=0,
    max_value=10000,
    value=10,
    help="Only show pages with at least this many clicks at peak"
)

st.sidebar.markdown("---")
st.sidebar.header("Excel Formatting")

highlight_color = st.sidebar.color_picker(
    "Peak month highlight color",
    "#00D1B1",
    help="Color to highlight the peak month"
)


def get_complete_months(df, months_to_analyze):
    """Identify complete months in the dataset."""
    df['month_date'] = df['date'].dt.to_period('M').astype(str)
    df['month_start'] = pd.to_datetime(df['month_date'] + '-01')
    df['month_end'] = df['month_start'] + pd.offsets.MonthEnd(0)

    monthly_coverage = df.groupby('month_date').agg({
        'date': lambda x: len(pd.date_range(start=x.min(), end=x.max(), freq='D')),
        'month_start': 'first',
        'month_end': 'first'
    })

    monthly_coverage['expected_days'] = monthly_coverage['month_end'].dt.day
    complete_months = monthly_coverage[
        monthly_coverage['date'] >= monthly_coverage['expected_days']
    ].index.tolist()

    complete_months.sort()
    return complete_months[-months_to_analyze:]


def create_excel_report(df, highlight_color):
    """Create formatted Excel report."""
    output = BytesIO()
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = 'Content Decay'

    # Excel style configuration
    hex_color = highlight_color.lstrip('#')
    green_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
    link_font = Font(color='0000FF', underline='single')

    # Write headers and data
    headers = ['Url'] + [col for col in df.columns if col != 'Url']
    ws_data.append(headers)

    for _, row in df.iterrows():
        ws_data.append([row[col] for col in headers])

    # Create table
    tab = Table(
        displayName="ContentDecayTable",
        ref=f"A1:{get_column_letter(ws_data.max_column)}{ws_data.max_row}",
        tableStyleInfo=TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    )
    ws_data.add_table(tab)

    # Format rows
    for row_idx, row in enumerate(ws_data.rows, 1):
        if row_idx == 1:
            continue

        # Format URL
        url_cell = row[0]
        if url_cell.value and isinstance(url_cell.value, str) and url_cell.value.startswith('http'):
            url_cell.font = link_font
            url_cell.hyperlink = url_cell.value

        # Highlight peak value (excluding URL and Clicks Lost columns)
        values = [(cell.value or 0, cell_idx) for cell_idx, cell in enumerate(row[1:-1], 1)]
        if values:
            peak_value, peak_col = max(values, key=lambda x: x[0])
            if peak_value > 0:
                peak_cell = row[peak_col]
                peak_cell.fill = green_fill
                peak_cell.font = Font(color="FFFFFF", bold=True)

    # Adjust column widths
    ws_data.column_dimensions['A'].width = 50
    for col in range(2, ws_data.max_column + 1):
        ws_data.column_dimensions[get_column_letter(col)].width = 12

    # Create Info Sheet
    ws_info = wb.create_sheet('Info')
    ws_info['A1'] = 'Content Decay Report'
    ws_info['A3'] = "Report Information"
    ws_info['A4'] = "Showing pages that have lost clicks compared to their peak month."
    ws_info['A5'] = "The highest number of clicks are shown in green."
    ws_info['A7'] = "Clicks Lost Explained"
    ws_info['A8'] = "The 'Clicks Lost' metric represents the decrease in clicks a page receives in the latest month compared to its peak performance."
    ws_info['A10'] = "How to Use"
    ws_info['A11'] = "1. Sort by 'Clicks Lost' to prioritize pages with the most significant decline"
    ws_info['A12'] = "2. Investigate potential causes: outdated content, algorithm changes, competition"
    ws_info['A13'] = "3. Update and optimize content to recover performance"

    ws_info['A1'].font = Font(bold=True, size=14)
    ws_info['A3'].font = Font(bold=True)
    ws_info['A7'].font = Font(bold=True)
    ws_info['A10'].font = Font(bold=True)
    ws_info.column_dimensions['A'].width = 100

    wb.save(output)
    return output.getvalue()


# File upload
st.subheader("Upload GSC Data")
gsc_file = st.file_uploader(
    "Upload CSV with GSC data",
    type=['csv'],
    help="CSV with date, page, query, clicks columns"
)

if gsc_file is not None:
    try:
        # Load data
        try:
            df = pd.read_csv(gsc_file, encoding='utf-8')
        except:
            gsc_file.seek(0)
            df = pd.read_csv(gsc_file, encoding='latin-1')

        st.success(f"Loaded {len(df):,} rows")

        # Column mapping
        with st.expander("Column Mapping"):
            cols = df.columns.tolist()

            date_col = st.selectbox(
                "Date column",
                cols,
                index=cols.index('date') if 'date' in cols else 0
            )
            page_col = st.selectbox(
                "Page/URL column",
                cols,
                index=cols.index('page') if 'page' in cols else 0
            )
            clicks_col = st.selectbox(
                "Clicks column",
                cols,
                index=cols.index('clicks') if 'clicks' in cols else 0
            )

        with st.expander("Preview data"):
            st.dataframe(df.head(20))

        if st.button("Analyze Content Decay", type="primary"):
            with st.spinner("Analyzing content decay patterns..."):
                df_work = df.copy()

                # Parse date
                df_work[date_col] = pd.to_datetime(df_work[date_col])

                # Get complete months
                complete_months = get_complete_months(df_work, months_to_analyze)

                if not complete_months:
                    st.error("No complete months found in the data")
                else:
                    st.info(f"Analyzing {len(complete_months)} complete months: {complete_months[0]} to {complete_months[-1]}")

                    # Filter to complete months
                    df_work['Year-Month'] = df_work[date_col].dt.strftime('%Y-%m')
                    df_recent = df_work[df_work['Year-Month'].isin(complete_months)].copy()

                    # Get all unique pages
                    all_pages = df_recent[page_col].unique()

                    # Aggregate clicks by page and month
                    monthly_clicks = df_recent.groupby([page_col, 'Year-Month'])[clicks_col].sum()

                    # Build result dataframe
                    result_df = pd.DataFrame(index=all_pages, columns=['Url'] + complete_months + ['Clicks Lost'])
                    result_df['Url'] = result_df.index

                    # Initialize with zeros
                    for month in complete_months:
                        result_df[month] = 0

                    # Fill in click data
                    progress_bar = st.progress(0)
                    for i, page in enumerate(all_pages):
                        for month in complete_months:
                            value = monthly_clicks.get((page, month), 0)
                            result_df.at[page, month] = int(value)
                        progress_bar.progress((i + 1) / len(all_pages))

                    # Calculate clicks lost
                    month_data = result_df[complete_months]
                    peak_clicks = month_data.max(axis=1)
                    latest_month_clicks = result_df[complete_months[-1]]
                    result_df['Clicks Lost'] = (latest_month_clicks - peak_clicks).astype(int)

                    # Filter to decaying content with sufficient traffic
                    result_df = result_df[result_df['Clicks Lost'] < 0]
                    result_df = result_df[peak_clicks >= min_peak_clicks]
                    result_df = result_df.sort_values('Clicks Lost')
                    result_df = result_df.reset_index(drop=True)

                    # Display results
                    st.subheader("Content Decay Results")

                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Decaying Pages", f"{len(result_df):,}")
                    with col2:
                        total_lost = result_df['Clicks Lost'].sum()
                        st.metric("Total Clicks Lost", f"{abs(total_lost):,}")
                    with col3:
                        avg_lost = result_df['Clicks Lost'].mean()
                        st.metric("Avg Clicks Lost", f"{abs(avg_lost):,.0f}")
                    with col4:
                        worst_page_loss = result_df['Clicks Lost'].min()
                        st.metric("Worst Page Loss", f"{abs(worst_page_loss):,}")

                    # Show table
                    st.subheader("Decaying Pages")
                    st.dataframe(result_df, use_container_width=True)

                    # Download
                    st.subheader("Download Report")

                    excel_data = create_excel_report(result_df, highlight_color)
                    st.download_button(
                        label="Download Excel Report",
                        data=excel_data,
                        file_name="content_decay_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # Also provide CSV
                    csv_data = result_df.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="Download CSV",
                        data=csv_data,
                        file_name="content_decay_report.csv",
                        mime="text/csv"
                    )

    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        import traceback
        st.code(traceback.format_exc())

else:
    st.info("Upload a GSC export CSV to begin")

    st.subheader("Required Data Format")
    example_data = {
        "date": ["2024-01-15", "2024-01-15", "2024-02-15", "2024-02-15"],
        "page": ["https://example.com/page1", "https://example.com/page2", "https://example.com/page1", "https://example.com/page2"],
        "query": ["keyword 1", "keyword 2", "keyword 1", "keyword 2"],
        "clicks": [150, 200, 80, 180]
    }
    st.dataframe(pd.DataFrame(example_data))

    st.subheader("Example Output")
    output_example = {
        "Url": ["https://example.com/page1", "https://example.com/page3"],
        "2024-01": [500, 300],
        "2024-02": [450, 200],
        "2024-03": [200, 150],
        "Clicks Lost": [-300, -150]
    }
    st.dataframe(pd.DataFrame(output_example))
