#!/usr/bin/env python3
"""
Content Decay Analyzer - CLI Version

Analyze GSC data to find pages losing traffic over time.

Usage:
    python content_decay_analyzer_cli.py --input gsc_data.csv

Author: Lee Foot
Website: https://www.leefoot.com
"""

import argparse
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def get_complete_months(df, date_col, months_to_analyze):
    """Identify complete months in the dataset."""
    df['month_date'] = df[date_col].dt.to_period('M').astype(str)
    df['month_start'] = pd.to_datetime(df['month_date'] + '-01')
    df['month_end'] = df['month_start'] + pd.offsets.MonthEnd(0)

    monthly_coverage = df.groupby('month_date').agg({
        date_col: lambda x: len(pd.date_range(start=x.min(), end=x.max(), freq='D')),
        'month_start': 'first',
        'month_end': 'first'
    })

    monthly_coverage['expected_days'] = monthly_coverage['month_end'].dt.day
    complete_months = monthly_coverage[
        monthly_coverage[date_col] >= monthly_coverage['expected_days']
    ].index.tolist()

    complete_months.sort()
    return complete_months[-months_to_analyze:]


def create_excel_report(df, output_path, highlight_color='00D1B1'):
    """Create formatted Excel report."""
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = 'Content Decay'

    green_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type='solid')
    link_font = Font(color='0000FF', underline='single')

    headers = ['Url'] + [col for col in df.columns if col != 'Url']
    ws_data.append(headers)

    for _, row in df.iterrows():
        ws_data.append([row[col] for col in headers])

    tab = Table(
        displayName="ContentDecayTable",
        ref=f"A1:{get_column_letter(ws_data.max_column)}{ws_data.max_row}",
        tableStyleInfo=TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    )
    ws_data.add_table(tab)

    for row_idx, row in enumerate(ws_data.rows, 1):
        if row_idx == 1:
            continue

        url_cell = row[0]
        if url_cell.value and isinstance(url_cell.value, str) and url_cell.value.startswith('http'):
            url_cell.font = link_font
            url_cell.hyperlink = url_cell.value

        values = [(cell.value or 0, cell_idx) for cell_idx, cell in enumerate(row[1:-1], 1)]
        if values:
            peak_value, peak_col = max(values, key=lambda x: x[0])
            if peak_value > 0:
                peak_cell = row[peak_col]
                peak_cell.fill = green_fill
                peak_cell.font = Font(color="FFFFFF", bold=True)

    ws_data.column_dimensions['A'].width = 50
    for col in range(2, ws_data.max_column + 1):
        ws_data.column_dimensions[get_column_letter(col)].width = 12

    # Info sheet
    ws_info = wb.create_sheet('Info')
    ws_info['A1'] = 'Content Decay Report'
    ws_info['A3'] = "Showing pages that have lost clicks compared to their peak month."
    ws_info['A5'] = "Clicks Lost = Latest Month Clicks - Peak Month Clicks"
    ws_info['A1'].font = Font(bold=True, size=14)
    ws_info.column_dimensions['A'].width = 80

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='Analyze GSC data to find pages losing traffic over time'
    )
    parser.add_argument('--input', required=True, help='Input CSV with GSC data')
    parser.add_argument('--output', default='content_decay_report.xlsx',
                        help='Output Excel path (default: content_decay_report.xlsx)')
    parser.add_argument('--months', type=int, default=12,
                        help='Number of months to analyze (default: 12)')
    parser.add_argument('--min-peak-clicks', type=int, default=10,
                        help='Minimum peak clicks threshold (default: 10)')
    parser.add_argument('--date-col', default='date', help='Date column name')
    parser.add_argument('--page-col', default='page', help='Page/URL column name')
    parser.add_argument('--clicks-col', default='clicks', help='Clicks column name')

    args = parser.parse_args()

    # Load data
    print(f"Loading: {args.input}")
    try:
        df = pd.read_csv(args.input, encoding='utf-8')
    except:
        df = pd.read_csv(args.input, encoding='latin-1')
    print(f"  Loaded {len(df):,} rows")

    # Parse date
    df[args.date_col] = pd.to_datetime(df[args.date_col])

    # Get complete months
    complete_months = get_complete_months(df.copy(), args.date_col, args.months)

    if not complete_months:
        print("Error: No complete months found in the data")
        return

    print(f"  Analyzing {len(complete_months)} complete months: {complete_months[0]} to {complete_months[-1]}")

    # Filter to complete months
    df['Year-Month'] = df[args.date_col].dt.strftime('%Y-%m')
    df_recent = df[df['Year-Month'].isin(complete_months)].copy()

    # Aggregate
    all_pages = df_recent[args.page_col].unique()
    monthly_clicks = df_recent.groupby([args.page_col, 'Year-Month'])[args.clicks_col].sum()

    print(f"  Found {len(all_pages):,} unique pages")

    # Build result dataframe
    result_df = pd.DataFrame(index=all_pages, columns=['Url'] + complete_months + ['Clicks Lost'])
    result_df['Url'] = result_df.index

    for month in complete_months:
        result_df[month] = 0

    for page in all_pages:
        for month in complete_months:
            value = monthly_clicks.get((page, month), 0)
            result_df.at[page, month] = int(value)

    # Calculate clicks lost
    month_data = result_df[complete_months]
    peak_clicks = month_data.max(axis=1)
    latest_month_clicks = result_df[complete_months[-1]]
    result_df['Clicks Lost'] = (latest_month_clicks - peak_clicks).astype(int)

    # Filter
    result_df = result_df[result_df['Clicks Lost'] < 0]
    result_df = result_df[peak_clicks >= args.min_peak_clicks]
    result_df = result_df.sort_values('Clicks Lost')
    result_df = result_df.reset_index(drop=True)

    # Save
    if args.output.endswith('.xlsx'):
        create_excel_report(result_df, args.output)
    else:
        result_df.to_csv(args.output, index=False, encoding='utf-8-sig')

    print(f"\nResults saved to: {args.output}")
    print(f"  Decaying pages: {len(result_df):,}")
    print(f"  Total clicks lost: {abs(result_df['Clicks Lost'].sum()):,}")

    # Show worst pages
    print(f"\nWorst decaying pages:")
    for _, row in result_df.head(10).iterrows():
        print(f"  [{row['Clicks Lost']:,}] {row['Url'][:60]}...")


if __name__ == '__main__':
    main()
