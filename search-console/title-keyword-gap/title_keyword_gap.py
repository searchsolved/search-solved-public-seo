# Author: Lee Foot
# Website: https://leefoot.com

####################################################################################
#                                                                                  #
#  Title Keyword Gap Finder v2                                                     #
#                                                                                  #
#  Two modes:                                                                      #
#  1. Keyword Gap - Find GSC keywords driving traffic but missing from titles.     #
#  2. Segment Analysis - Split titles by delimiter, surface segments with no       #
#     search volume and GSC keywords not represented in any title segment.         #
#                                                                                  #
####################################################################################
# Author   : Lee Foot                                                              #
# Website  : https://leefoot.com                                                   #
# Contact  : https://www.leefoot.com/contact                                       #
# Email    : hello@leefoot.com                                                     #
# LinkedIn : https://www.linkedin.com/in/lee-foot/                                 #
# Bluesky  : https://bsky.app/profile/leefootseo.bsky.social                       #
####################################################################################

"""
Title Keyword Gap Finder v2

Two analysis modes for page title optimisation:

Mode 1 - Keyword Gap:
    Compares Google Search Console keywords against page titles to identify
    keywords that are driving impressions but are missing from the page title.
    Great for quick-win title optimisation opportunities.

Mode 2 - Title Segment Analysis:
    Splits page titles by a delimiter (e.g. "|"), treats each segment as a
    potential keyword, then cross-references against GSC query data. Surfaces:
    - Title segments with zero search volume (wasted title real estate)
    - High-performing GSC keywords not represented in any title segment
    Output is a highlighted Excel file showing matches and gaps.
"""

import streamlit as st
import pandas as pd
from io import BytesIO

st.set_page_config(page_title="Title Keyword Gap Finder v2", page_icon="🔎", layout="wide")

st.title("Title Keyword Gap Finder v2")
st.markdown(
    "*Created by* "
    "[![Website](https://img.shields.io/badge/-leefoot.com-2A9D8F?logoColor=white)]"
    "(https://www.leefoot.com) · "
    "[![Hire Me](https://img.shields.io/badge/-Hire%20Me-FF6B6B?logoColor=white)]"
    "(https://www.leefoot.com/contact) · "
    "[![LinkedIn](https://img.shields.io/badge/-LinkedIn-0A66C2?logo=linkedin&logoColor=white)]"
    "(https://www.linkedin.com/in/lee-foot/) · "
    "[![Bluesky](https://img.shields.io/badge/-Bluesky-0285FF?logoColor=white)]"
    "(https://bsky.app/profile/leefootseo.bsky.social) · "
    "[![More Tools](https://img.shields.io/badge/-More%20Tools-8B5CF6?logoColor=white)]"
    "(https://leefoot.com/tools) · "
    "[![GitHub](https://img.shields.io/badge/-GitHub-6B7280?logoColor=white)]"
    "(https://github.com/searchsolved/search-solved-public-seo)"
)

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def load_csv(file):
    """Load CSV with encoding fallback."""
    try:
        return pd.read_csv(file, encoding="utf-8")
    except Exception:
        file.seek(0)
        return pd.read_csv(file, encoding="latin-1")


def find_column(df, possible_names):
    """Find a column by trying multiple possible names (case-insensitive)."""
    for name in possible_names:
        for col in df.columns:
            if name.lower() == col.lower() or name.lower() in col.lower():
                return col
    return None


# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------

tab1, tab2 = st.tabs(["Keyword Gap", "Title Segment Analysis"])

# ===========================================================================
# TAB 1 - KEYWORD GAP (original mode)
# ===========================================================================

with tab1:
    with st.expander("How to use this mode"):
        st.markdown("""
        **What this does:**
        - Compares your Google Search Console queries against page titles
        - Finds keywords that drive impressions but are not in the title
        - Highlights which keywords are already in your titles
        - Suggests title optimisation opportunities

        **How to get the data:**

        **1. Screaming Frog Crawl:**
        - Crawl your site with Screaming Frog
        - Export `internal_html.csv` from the Internal tab
        - Required columns: Address, Title 1

        **2. GSC Query Data:**
        - Go to Search Console > Performance > Pages
        - Export the data (clicks, impressions, queries by page)
        - Or use the GSC API to export query-level data

        **Required GSC columns:**
        - `page` - The landing page URL
        - `query` - The search query
        - `clicks` - Click count
        - `impressions` - Impression count
        """)

    # Sidebar settings (shared)
    st.sidebar.header("Settings")

    title_delimiter = st.sidebar.text_input(
        "Title delimiter",
        value="|",
        help="Character used to split brand from title (e.g., | or -)",
    )

    branding = st.sidebar.text_input(
        "Brand terms to exclude",
        value="",
        help="Brand name(s) to filter out of analysis (comma-separated)",
    )

    url_filter = st.sidebar.text_input(
        "URL filter (optional)",
        value="",
        help="Only analyse URLs containing this text (e.g., /products/)",
    )

    max_keywords_per_page = st.sidebar.number_input(
        "Max keywords per page",
        min_value=5,
        max_value=50,
        value=10,
        help="Maximum GSC keywords to show per page",
    )

    min_impressions = st.sidebar.number_input(
        "Minimum impressions",
        min_value=0,
        max_value=10000,
        value=0,
        help="Only include queries with at least this many impressions",
    )

    # File uploads for Tab 1
    st.subheader("Upload Files")
    col1, col2 = st.columns(2)

    with col1:
        crawl_file = st.file_uploader(
            "Screaming Frog crawl (CSV)",
            type=["csv"],
            key="crawl_tab1",
            help="Export from Screaming Frog with titles",
        )

    with col2:
        gsc_file = st.file_uploader(
            "GSC query data (CSV)",
            type=["csv"],
            key="gsc_tab1",
            help="GSC export with page, query, clicks, impressions",
        )

    if crawl_file is not None and gsc_file is not None:
        try:
            df_crawl = load_csv(crawl_file)
            df_gsc = load_csv(gsc_file)

            st.success(f"Loaded crawl: {len(df_crawl):,} URLs | GSC: {len(df_gsc):,} queries")

            # Find columns in crawl
            address_col = find_column(df_crawl, ["address", "url"])
            title_col = find_column(df_crawl, ["title 1", "title", "page title"])

            # Find columns in GSC
            page_col = find_column(df_gsc, ["page", "landing page", "url"])
            query_col = find_column(df_gsc, ["query", "keyword", "top queries"])
            clicks_col = find_column(df_gsc, ["clicks", "click"])
            impressions_col = find_column(df_gsc, ["impressions", "impression"])

            with st.expander("Column Mapping"):
                st.markdown("**Crawl columns:**")
                c1, c2 = st.columns(2)
                with c1:
                    address_col = st.selectbox(
                        "URL column",
                        df_crawl.columns.tolist(),
                        index=df_crawl.columns.tolist().index(address_col)
                        if address_col
                        else 0,
                        key="t1_address",
                    )
                with c2:
                    title_col = st.selectbox(
                        "Title column",
                        df_crawl.columns.tolist(),
                        index=df_crawl.columns.tolist().index(title_col)
                        if title_col
                        else 0,
                        key="t1_title",
                    )

                st.markdown("**GSC columns:**")
                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    page_col = st.selectbox(
                        "Page column",
                        df_gsc.columns.tolist(),
                        index=df_gsc.columns.tolist().index(page_col) if page_col else 0,
                        key="t1_page",
                    )
                with c2:
                    query_col = st.selectbox(
                        "Query column",
                        df_gsc.columns.tolist(),
                        index=df_gsc.columns.tolist().index(query_col)
                        if query_col
                        else 0,
                        key="t1_query",
                    )
                with c3:
                    clicks_col = st.selectbox(
                        "Clicks column",
                        df_gsc.columns.tolist(),
                        index=df_gsc.columns.tolist().index(clicks_col)
                        if clicks_col
                        else 0,
                        key="t1_clicks",
                    )
                with c4:
                    impressions_col = st.selectbox(
                        "Impressions column",
                        df_gsc.columns.tolist(),
                        index=df_gsc.columns.tolist().index(impressions_col)
                        if impressions_col
                        else 0,
                        key="t1_impressions",
                    )

            if st.button("Analyse Title Gaps", type="primary", key="btn_tab1"):
                with st.spinner("Analysing keywords vs titles..."):
                    # Prepare crawl data
                    df_titles = df_crawl[[address_col, title_col]].copy()
                    df_titles.columns = ["page", "title"]
                    df_titles = df_titles.dropna(subset=["title"])

                    if url_filter.strip():
                        df_titles = df_titles[
                            df_titles["page"].str.contains(url_filter, na=False)
                        ]

                    # Prepare GSC data
                    df_queries = df_gsc[
                        [page_col, query_col, clicks_col, impressions_col]
                    ].copy()
                    df_queries.columns = ["page", "query", "clicks", "impressions"]

                    if min_impressions > 0:
                        df_queries = df_queries[
                            df_queries["impressions"] >= min_impressions
                        ]

                    if url_filter.strip():
                        df_queries = df_queries[
                            df_queries["page"].str.contains(url_filter, na=False)
                        ]

                    # Filter out brand terms
                    if branding.strip():
                        brand_terms = [
                            b.strip().lower() for b in branding.split(",") if b.strip()
                        ]
                        for term in brand_terms:
                            df_queries = df_queries[
                                ~df_queries["query"]
                                .str.lower()
                                .str.contains(term, na=False)
                            ]

                    # Sort and limit keywords per page
                    df_queries = df_queries.sort_values(
                        ["page", "clicks"], ascending=[True, False]
                    )
                    df_queries = df_queries.groupby("page").head(max_keywords_per_page)

                    # Merge with titles
                    df_merged = pd.merge(df_queries, df_titles, on="page", how="inner")

                    if len(df_merged) == 0:
                        st.warning(
                            "No matching pages found between crawl and GSC data. "
                            "Check that URLs match exactly."
                        )
                    else:

                        def check_query_in_title(row):
                            query = str(row["query"]).strip().lower()
                            title = str(row["title"]).strip().lower()
                            if title_delimiter:
                                title_parts = [
                                    p.strip() for p in title.split(title_delimiter)
                                ]
                            else:
                                title_parts = [title]
                            for part in title_parts:
                                if query in part:
                                    return True
                            return False

                        df_merged["in_title"] = df_merged.apply(
                            check_query_in_title, axis=1
                        )

                        df_merged["total_clicks"] = df_merged.groupby("page")[
                            "clicks"
                        ].transform("sum")
                        df_merged["total_impressions"] = df_merged.groupby("page")[
                            "impressions"
                        ].transform("sum")

                        df_merged = df_merged.sort_values(
                            by=["total_impressions", "page", "clicks"],
                            ascending=[False, True, False],
                        )

                        # Display results
                        st.subheader("Results")

                        c1, c2, c3, c4 = st.columns(4)
                        with c1:
                            st.metric(
                                "Pages Analysed", f"{df_merged['page'].nunique():,}"
                            )
                        with c2:
                            st.metric("Keywords Analysed", f"{len(df_merged):,}")
                        with c3:
                            in_title_count = df_merged["in_title"].sum()
                            st.metric("Already in Title", f"{in_title_count:,}")
                        with c4:
                            not_in_title_count = len(df_merged) - in_title_count
                            st.metric("Missing from Title", f"{not_in_title_count:,}")

                        # Show gaps
                        st.subheader("Keywords Missing from Titles (Opportunities)")
                        df_gaps = df_merged[~df_merged["in_title"]].copy()
                        df_gaps = df_gaps.sort_values("impressions", ascending=False)

                        if len(df_gaps) > 0:
                            st.dataframe(
                                df_gaps[
                                    [
                                        "page",
                                        "title",
                                        "query",
                                        "clicks",
                                        "impressions",
                                    ]
                                ].head(100),
                                use_container_width=True,
                            )

                            st.subheader("Most Common Missing Keywords")
                            missing_kws = (
                                df_gaps["query"]
                                .value_counts()
                                .head(20)
                                .reset_index()
                            )
                            missing_kws.columns = ["Keyword", "Pages Missing It"]
                            c1, c2 = st.columns(2)
                            with c1:
                                st.dataframe(missing_kws, use_container_width=True)
                            with c2:
                                st.bar_chart(
                                    missing_kws.head(15).set_index("Keyword")
                                )
                        else:
                            st.success(
                                "All analysed keywords are already present in page titles!"
                            )

                        # Download options
                        st.subheader("Download")

                        c1, c2 = st.columns(2)

                        with c1:
                            csv_full = df_merged.to_csv(index=False).encode("utf-8-sig")
                            st.download_button(
                                label="Download All Results (CSV)",
                                data=csv_full,
                                file_name="title_keyword_analysis.csv",
                                mime="text/csv",
                                key="dl_csv_full_t1",
                            )

                        with c2:
                            csv_gaps = df_gaps.to_csv(index=False).encode("utf-8-sig")
                            st.download_button(
                                label="Download Gaps Only (CSV)",
                                data=csv_gaps,
                                file_name="title_keyword_gaps.csv",
                                mime="text/csv",
                                key="dl_csv_gaps_t1",
                            )

                        # Excel with highlighting
                        st.markdown("---")
                        try:
                            from openpyxl.styles import PatternFill

                            output = BytesIO()
                            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                                df_merged.to_excel(
                                    writer, index=False, sheet_name="Analysis"
                                )
                                ws = writer.sheets["Analysis"]

                                green_fill = PatternFill(
                                    start_color="90EE90",
                                    end_color="90EE90",
                                    fill_type="solid",
                                )
                                yellow_fill = PatternFill(
                                    start_color="FFFF00",
                                    end_color="FFFF00",
                                    fill_type="solid",
                                )

                                in_title_col_idx = None
                                for idx, cell in enumerate(ws[1], 1):
                                    if cell.value == "in_title":
                                        in_title_col_idx = idx
                                        break

                                if in_title_col_idx:
                                    for row_idx in range(2, ws.max_row + 1):
                                        cell_value = ws.cell(
                                            row=row_idx, column=in_title_col_idx
                                        ).value
                                        fill = (
                                            green_fill if cell_value else yellow_fill
                                        )
                                        for col_idx in range(1, ws.max_column + 1):
                                            ws.cell(
                                                row=row_idx, column=col_idx
                                            ).fill = fill

                            excel_data = output.getvalue()
                            st.download_button(
                                label="Download Excel with Highlighting",
                                data=excel_data,
                                file_name="title_keyword_analysis.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_excel_t1",
                            )
                        except ImportError:
                            st.info(
                                "Install openpyxl for Excel export with highlighting"
                            )

        except Exception as e:
            st.error(f"Error processing files: {str(e)}")
            import traceback

            st.code(traceback.format_exc())

    else:
        st.info("Upload both a crawl file and GSC data to begin")

        st.subheader("Example Output")
        example_data = {
            "Page": ["/products/widget", "/products/widget", "/products/gadget"],
            "Title": [
                "Best Widget for Home",
                "Best Widget for Home",
                "Professional Gadget Tool",
            ],
            "Query": ["widget", "industrial widget", "gadget"],
            "Impressions": [1500, 800, 2000],
            "In Title": [True, False, True],
        }
        st.dataframe(pd.DataFrame(example_data))
        st.markdown("*Yellow highlight = keyword missing from title (opportunity)*")


# ===========================================================================
# TAB 2 - TITLE SEGMENT ANALYSIS
# ===========================================================================

with tab2:
    with st.expander("How to use this mode"):
        st.markdown("""
        **What this does:**
        - Splits each page title by a delimiter (e.g. `|`) into segments
        - Treats each segment as a potential keyword
        - Cross-references segments against your GSC queries for that page
        - Highlights which title segments have search volume (green)
        - Surfaces segments with no GSC impressions (wasted title space)
        - Shows top GSC keywords that are not represented in any title segment

        **Use case:**
        Many category or listing pages use delimiter-separated titles like:
        `Widgets | Industrial Supplies | Example Store`

        This mode tells you whether each segment is pulling its weight in
        search, and which proven keywords could replace underperforming ones.

        **Data required:**

        **1. Crawl CSV** (Screaming Frog or similar):
        - Columns: Address (URL), Title 1

        **2. GSC Query CSV** (Search Console export):
        - Columns: page, query, clicks, impressions
        """)

    st.subheader("Upload Files")
    col1, col2 = st.columns(2)

    with col1:
        crawl_file_t2 = st.file_uploader(
            "Crawl CSV (Address + Title 1)",
            type=["csv"],
            key="crawl_tab2",
            help="Screaming Frog or similar crawl export",
        )

    with col2:
        gsc_file_t2 = st.file_uploader(
            "GSC query data (CSV)",
            type=["csv"],
            key="gsc_tab2",
            help="GSC export with page, query, clicks, impressions",
        )

    if crawl_file_t2 is not None and gsc_file_t2 is not None:
        try:
            df_crawl_t2 = load_csv(crawl_file_t2)
            df_gsc_t2 = load_csv(gsc_file_t2)

            st.success(
                f"Loaded crawl: {len(df_crawl_t2):,} URLs | "
                f"GSC: {len(df_gsc_t2):,} queries"
            )

            # Find columns
            address_col_t2 = find_column(df_crawl_t2, ["address", "url"])
            title_col_t2 = find_column(df_crawl_t2, ["title 1", "title", "page title"])

            page_col_t2 = find_column(df_gsc_t2, ["page", "landing page", "url"])
            query_col_t2 = find_column(df_gsc_t2, ["query", "keyword", "top queries"])
            clicks_col_t2 = find_column(df_gsc_t2, ["clicks", "click"])
            impressions_col_t2 = find_column(
                df_gsc_t2, ["impressions", "impression"]
            )

            with st.expander("Column Mapping"):
                st.markdown("**Crawl columns:**")
                c1, c2 = st.columns(2)
                with c1:
                    address_col_t2 = st.selectbox(
                        "URL column",
                        df_crawl_t2.columns.tolist(),
                        index=df_crawl_t2.columns.tolist().index(address_col_t2)
                        if address_col_t2
                        else 0,
                        key="t2_address",
                    )
                with c2:
                    title_col_t2 = st.selectbox(
                        "Title column",
                        df_crawl_t2.columns.tolist(),
                        index=df_crawl_t2.columns.tolist().index(title_col_t2)
                        if title_col_t2
                        else 0,
                        key="t2_title",
                    )

                st.markdown("**GSC columns:**")
                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    page_col_t2 = st.selectbox(
                        "Page column",
                        df_gsc_t2.columns.tolist(),
                        index=df_gsc_t2.columns.tolist().index(page_col_t2)
                        if page_col_t2
                        else 0,
                        key="t2_page",
                    )
                with c2:
                    query_col_t2 = st.selectbox(
                        "Query column",
                        df_gsc_t2.columns.tolist(),
                        index=df_gsc_t2.columns.tolist().index(query_col_t2)
                        if query_col_t2
                        else 0,
                        key="t2_query",
                    )
                with c3:
                    clicks_col_t2 = st.selectbox(
                        "Clicks column",
                        df_gsc_t2.columns.tolist(),
                        index=df_gsc_t2.columns.tolist().index(clicks_col_t2)
                        if clicks_col_t2
                        else 0,
                        key="t2_clicks",
                    )
                with c4:
                    impressions_col_t2 = st.selectbox(
                        "Impressions column",
                        df_gsc_t2.columns.tolist(),
                        index=df_gsc_t2.columns.tolist().index(impressions_col_t2)
                        if impressions_col_t2
                        else 0,
                        key="t2_impressions",
                    )

            if st.button("Analyse Title Segments", type="primary", key="btn_tab2"):
                with st.spinner("Analysing title segments vs GSC data..."):

                    # -- Prepare crawl data --
                    df_titles_t2 = df_crawl_t2[
                        [address_col_t2, title_col_t2]
                    ].copy()
                    df_titles_t2.columns = ["page", "title"]
                    df_titles_t2 = df_titles_t2.dropna(subset=["page", "title"])

                    if url_filter.strip():
                        df_titles_t2 = df_titles_t2[
                            df_titles_t2["page"].str.contains(url_filter, na=False)
                        ]

                    # -- Prepare GSC data --
                    df_gsc_prep = df_gsc_t2[
                        [page_col_t2, query_col_t2, clicks_col_t2, impressions_col_t2]
                    ].copy()
                    df_gsc_prep.columns = ["page", "query", "clicks", "impressions"]

                    # Filter out brand terms
                    if branding.strip():
                        brand_terms = [
                            b.strip().lower()
                            for b in branding.split(",")
                            if b.strip()
                        ]
                        for term in brand_terms:
                            df_gsc_prep = df_gsc_prep[
                                ~df_gsc_prep["query"]
                                .str.lower()
                                .str.contains(term, na=False)
                            ]

                    if url_filter.strip():
                        df_gsc_prep = df_gsc_prep[
                            df_gsc_prep["page"].str.contains(url_filter, na=False)
                        ]

                    if min_impressions > 0:
                        df_gsc_prep = df_gsc_prep[
                            df_gsc_prep["impressions"] >= min_impressions
                        ]

                    df_gsc_prep["kw_source"] = "gsc"

                    # -- Split titles into segments --
                    df_segments = df_titles_t2.copy()
                    df_segments = df_segments.join(
                        df_segments["title"]
                        .str.split(title_delimiter, expand=True)
                        .add_prefix("title_")
                    )

                    # Melt segments into rows
                    segment_cols = [
                        c
                        for c in df_segments.columns
                        if c.startswith("title_")
                    ]
                    df_segments["query"] = df_segments[segment_cols].values.tolist()
                    df_segments = df_segments[["page", "title", "query"]].explode(
                        "query"
                    )

                    # Clean up segment queries
                    df_segments["query"] = (
                        df_segments["query"].astype(str).str.strip().str.lower()
                    )
                    df_segments["query"] = (
                        df_segments["query"].str.split().str.join(" ")
                    )

                    # Remove rows that are the full title or the URL itself
                    df_segments["title_lower"] = df_segments["title"].str.lower()
                    df_segments = df_segments[
                        df_segments["query"] != df_segments["title_lower"]
                    ]
                    df_segments = df_segments[
                        df_segments["query"] != df_segments["page"]
                    ]
                    df_segments = df_segments.drop(columns=["title_lower"])

                    # Remove empty/none values
                    df_segments = df_segments[df_segments["query"].notna()]
                    df_segments = df_segments[df_segments["query"] != ""]
                    df_segments = df_segments[df_segments["query"] != "none"]
                    df_segments = df_segments[df_segments["query"] != "nan"]

                    # Filter out brand terms from segments
                    if branding.strip():
                        for term in brand_terms:
                            df_segments = df_segments[
                                ~df_segments["query"].str.contains(term, na=False)
                            ]

                    df_segments["kw_source"] = "page_title"

                    # -- Merge segment keywords with GSC data --
                    df_seg_merged = pd.merge(
                        df_segments,
                        df_gsc_prep[["query", "page", "clicks", "impressions"]],
                        on=["query", "page"],
                        how="left",
                    )
                    cols_order = [
                        "page",
                        "query",
                        "kw_source",
                        "clicks",
                        "impressions",
                    ]
                    df_seg_merged = df_seg_merged.reindex(columns=cols_order)

                    # -- Get top GSC keywords not in title segments --
                    df_gsc_top = df_gsc_prep.copy()
                    if url_filter.strip():
                        df_gsc_top = df_gsc_top[
                            df_gsc_top["page"].str.contains(url_filter, na=False)
                        ]
                    df_gsc_top = df_gsc_top.sort_values("clicks", ascending=False)
                    df_gsc_top = df_gsc_top[df_gsc_top["clicks"] > 0]
                    df_gsc_top = df_gsc_top.groupby("page").head(
                        max_keywords_per_page
                    )

                    # -- Combine segment keywords and top GSC keywords --
                    df_combined = pd.concat(
                        [df_seg_merged, df_gsc_top[cols_order]], ignore_index=True
                    )
                    df_combined.fillna(
                        {"clicks": 0, "impressions": 0}, inplace=True
                    )
                    df_combined.drop_duplicates(
                        subset=["query", "page"], keep="first", inplace=True
                    )
                    df_combined.sort_values(
                        ["page", "clicks"], ascending=[True, False], inplace=True
                    )

                    # -- Re-attach title --
                    df_combined = pd.merge(
                        df_combined,
                        df_titles_t2[["page", "title"]],
                        on="page",
                        how="left",
                    )
                    df_combined.drop_duplicates(
                        subset=["page", "query"], keep="first", inplace=True
                    )
                    df_combined = df_combined[df_combined["title"].notna()]

                    # -- Aggregates --
                    df_combined["total_clicks"] = df_combined.groupby("page")[
                        "clicks"
                    ].transform("sum")
                    df_combined["total_impressions"] = df_combined.groupby("page")[
                        "impressions"
                    ].transform("sum")
                    df_combined.sort_values(
                        by=["total_impressions", "page"],
                        ascending=[False, True],
                        inplace=True,
                    )

                    # -- Determine match status --
                    def check_match(row):
                        query = str(row["query"]).strip().lower()
                        title = str(row["title"]).strip().lower()
                        segments = [
                            s.strip().lower()
                            for s in title.split(title_delimiter)
                        ]
                        return any(query in seg for seg in segments)

                    df_combined["in_title"] = df_combined.apply(check_match, axis=1)

                    if len(df_combined) == 0:
                        st.warning("No data after processing. Check URL filters and column mappings.")
                    else:
                        # -- Display results --
                        st.subheader("Results")

                        c1, c2, c3, c4 = st.columns(4)
                        with c1:
                            st.metric(
                                "Pages Analysed",
                                f"{df_combined['page'].nunique():,}",
                            )
                        with c2:
                            st.metric(
                                "Keywords Analysed", f"{len(df_combined):,}"
                            )
                        with c3:
                            seg_count = (
                                df_combined["kw_source"] == "page_title"
                            ).sum()
                            st.metric("Title Segments", f"{seg_count:,}")
                        with c4:
                            gsc_only = (
                                (df_combined["kw_source"] == "gsc")
                                & (~df_combined["in_title"])
                            ).sum()
                            st.metric(
                                "GSC Keywords Not in Title", f"{gsc_only:,}"
                            )

                        # Segments with no search volume
                        st.subheader(
                            "Title Segments with No Search Volume (Wasted Space)"
                        )
                        df_no_vol = df_combined[
                            (df_combined["kw_source"] == "page_title")
                            & (df_combined["impressions"] == 0)
                        ].copy()

                        if len(df_no_vol) > 0:
                            st.dataframe(
                                df_no_vol[["page", "title", "query"]].head(100),
                                use_container_width=True,
                            )
                            st.caption(
                                f"{len(df_no_vol):,} title segments have zero "
                                f"GSC impressions for their page."
                            )
                        else:
                            st.success(
                                "All title segments have some search volume."
                            )

                        # GSC keywords not in title
                        st.subheader(
                            "High-Performing GSC Keywords Missing from Title"
                        )
                        df_missing = df_combined[
                            (df_combined["kw_source"] == "gsc")
                            & (~df_combined["in_title"])
                        ].copy()
                        df_missing = df_missing.sort_values(
                            "clicks", ascending=False
                        )

                        if len(df_missing) > 0:
                            st.dataframe(
                                df_missing[
                                    [
                                        "page",
                                        "title",
                                        "query",
                                        "clicks",
                                        "impressions",
                                    ]
                                ].head(100),
                                use_container_width=True,
                            )
                        else:
                            st.success(
                                "All top GSC keywords are represented in title segments."
                            )

                        # -- Downloads --
                        st.subheader("Download")

                        c1, c2 = st.columns(2)

                        with c1:
                            csv_out = df_combined.to_csv(index=False).encode(
                                "utf-8-sig"
                            )
                            st.download_button(
                                label="Download Full Analysis (CSV)",
                                data=csv_out,
                                file_name="title_segment_analysis.csv",
                                mime="text/csv",
                                key="dl_csv_t2",
                            )

                        with c2:
                            # Excel with highlighting
                            try:
                                from openpyxl.styles import PatternFill

                                output_t2 = BytesIO()
                                with pd.ExcelWriter(
                                    output_t2, engine="openpyxl"
                                ) as writer:
                                    df_combined.to_excel(
                                        writer,
                                        index=False,
                                        sheet_name="Segment Analysis",
                                    )
                                    ws = writer.sheets["Segment Analysis"]

                                    green_fill = PatternFill(
                                        start_color="90EE90",
                                        end_color="90EE90",
                                        fill_type="solid",
                                    )
                                    yellow_fill = PatternFill(
                                        start_color="FFFF00",
                                        end_color="FFFF00",
                                        fill_type="solid",
                                    )

                                    in_title_col_idx = None
                                    for idx, cell in enumerate(ws[1], 1):
                                        if cell.value == "in_title":
                                            in_title_col_idx = idx
                                            break

                                    if in_title_col_idx:
                                        for row_idx in range(2, ws.max_row + 1):
                                            cell_value = ws.cell(
                                                row=row_idx,
                                                column=in_title_col_idx,
                                            ).value
                                            fill = (
                                                green_fill
                                                if cell_value
                                                else yellow_fill
                                            )
                                            for col_idx in range(
                                                1, ws.max_column + 1
                                            ):
                                                ws.cell(
                                                    row=row_idx, column=col_idx
                                                ).fill = fill

                                excel_data_t2 = output_t2.getvalue()
                                st.download_button(
                                    label="Download Excel with Highlighting",
                                    data=excel_data_t2,
                                    file_name="title_segment_analysis.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_excel_t2",
                                )
                            except ImportError:
                                st.info(
                                    "Install openpyxl for Excel export with "
                                    "highlighting"
                                )

        except Exception as e:
            st.error(f"Error processing files: {str(e)}")
            import traceback

            st.code(traceback.format_exc())

    else:
        st.info("Upload both a crawl file and GSC data to begin")

        st.subheader("Example Output")
        example_data = {
            "Page": [
                "https://example.com/widgets",
                "https://example.com/widgets",
                "https://example.com/widgets",
            ],
            "Title": [
                "Widgets | Industrial Tools | Example Store",
                "Widgets | Industrial Tools | Example Store",
                "Widgets | Industrial Tools | Example Store",
            ],
            "Query": ["widgets", "industrial tools", "best widget prices"],
            "Source": ["page_title", "page_title", "gsc"],
            "Clicks": [120, 0, 85],
            "In Title": [True, True, False],
        }
        st.dataframe(pd.DataFrame(example_data))
        st.markdown(
            "*Green = segment with search volume. "
            "Yellow = segment with no volume or GSC keyword not in title.*"
        )
