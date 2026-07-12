# Author: Lee Foot
# Website: https://leefoot.com
"""
Title Keyword Gap Finder v2 - CLI Version

Two modes:
  keyword-gap   - Find GSC keywords driving traffic but missing from page titles.
  segment       - Split titles by delimiter, surface segments with no search volume
                  and GSC keywords not represented in any title segment.

Usage:
    python title_keyword_gap_cli.py keyword-gap --crawl crawl.csv --gsc gsc.csv
    python title_keyword_gap_cli.py segment --crawl crawl.csv --gsc gsc.csv

Author: Lee Foot
Website: https://leefoot.com
"""

import argparse
import sys

import pandas as pd

try:
    from openpyxl.styles import PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def load_csv(filepath):
    """Load CSV with encoding fallback."""
    try:
        return pd.read_csv(filepath, encoding="utf-8")
    except Exception:
        return pd.read_csv(filepath, encoding="latin-1")


def find_column(df, possible_names):
    """Find a column by trying multiple possible names (case-insensitive)."""
    for name in possible_names:
        for col in df.columns:
            if name.lower() == col.lower() or name.lower() in col.lower():
                return col
    return None


def save_excel_highlighted(df, output_path, sheet_name="Analysis"):
    """Save DataFrame to Excel with row highlighting based on in_title column."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        green_fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        in_title_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "in_title":
                in_title_col = idx
                break

        if in_title_col:
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=in_title_col).value
                fill = green_fill if cell_value else yellow_fill
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    print(f"Excel saved to: {output_path}")


def prepare_inputs(args):
    """Load and validate crawl + GSC data, returning prepared DataFrames."""
    print(f"Loading crawl: {args.crawl}")
    df_crawl = load_csv(args.crawl)
    print(f"  Loaded {len(df_crawl):,} URLs")

    print(f"Loading GSC data: {args.gsc}")
    df_gsc = load_csv(args.gsc)
    print(f"  Loaded {len(df_gsc):,} queries")

    # Find columns in crawl
    address_col = find_column(df_crawl, ["address", "url"])
    title_col = find_column(df_crawl, ["title 1", "title", "page title"])

    # Find columns in GSC
    page_col = find_column(df_gsc, ["page", "landing page", "url"])
    query_col = find_column(df_gsc, ["query", "keyword", "top queries"])
    clicks_col = find_column(df_gsc, ["clicks", "click"])
    impressions_col = find_column(df_gsc, ["impressions", "impression"])

    if not all([address_col, title_col, page_col, query_col, clicks_col, impressions_col]):
        print("Error: Could not find all required columns")
        print(f"  Crawl: address={address_col}, title={title_col}")
        print(
            f"  GSC: page={page_col}, query={query_col}, "
            f"clicks={clicks_col}, impressions={impressions_col}"
        )
        sys.exit(1)

    print("  Using columns:")
    print(f"    Crawl: address={address_col}, title={title_col}")
    print(f"    GSC: page={page_col}, query={query_col}")

    # Prepare crawl data
    df_titles = df_crawl[[address_col, title_col]].copy()
    df_titles.columns = ["page", "title"]
    df_titles = df_titles.dropna(subset=["title"])

    if args.url_filter:
        df_titles = df_titles[df_titles["page"].str.contains(args.url_filter, na=False)]
        print(f"  Filtered to {len(df_titles):,} URLs matching '{args.url_filter}'")

    # Prepare GSC data
    df_queries = df_gsc[[page_col, query_col, clicks_col, impressions_col]].copy()
    df_queries.columns = ["page", "query", "clicks", "impressions"]

    if args.min_impressions > 0:
        df_queries = df_queries[df_queries["impressions"] >= args.min_impressions]
        print(
            f"  Filtered to {len(df_queries):,} queries with "
            f">= {args.min_impressions} impressions"
        )

    if args.url_filter:
        df_queries = df_queries[
            df_queries["page"].str.contains(args.url_filter, na=False)
        ]

    # Filter out brand terms
    if args.brand:
        brand_terms = [b.strip().lower() for b in args.brand.split(",") if b.strip()]
        for term in brand_terms:
            df_queries = df_queries[
                ~df_queries["query"].str.lower().str.contains(term, na=False)
            ]
        print(f"  Filtered out brand terms: {brand_terms}")

    return df_titles, df_queries


# ---------------------------------------------------------------------------
# Mode 1: Keyword Gap
# ---------------------------------------------------------------------------


def run_keyword_gap(args):
    """Original keyword gap mode: find GSC keywords missing from titles."""
    df_titles, df_queries = prepare_inputs(args)

    # Sort and limit keywords per page
    df_queries = df_queries.sort_values(["page", "clicks"], ascending=[True, False])
    df_queries = df_queries.groupby("page").head(args.max_keywords)

    # Merge with titles
    df_merged = pd.merge(df_queries, df_titles, on="page", how="inner")

    if len(df_merged) == 0:
        print("No matching pages found between crawl and GSC data.")
        sys.exit(0)

    print(f"\nAnalysing {len(df_merged):,} keyword-page combinations...")

    # Check if query is in title
    def check_query_in_title(row):
        query = str(row["query"]).strip().lower()
        title = str(row["title"]).strip().lower()
        if args.delimiter:
            title_parts = [p.strip() for p in title.split(args.delimiter)]
        else:
            title_parts = [title]
        for part in title_parts:
            if query in part:
                return True
        return False

    df_merged["in_title"] = df_merged.apply(check_query_in_title, axis=1)

    # Calculate totals per page
    df_merged["total_clicks"] = df_merged.groupby("page")["clicks"].transform("sum")
    df_merged["total_impressions"] = df_merged.groupby("page")[
        "impressions"
    ].transform("sum")

    # Sort by potential
    df_merged = df_merged.sort_values(
        by=["total_impressions", "page", "clicks"],
        ascending=[False, True, False],
    )

    # Summary stats
    in_title = df_merged["in_title"].sum()
    not_in_title = len(df_merged) - in_title

    print(f"\nResults:")
    print(f"  Pages analysed: {df_merged['page'].nunique():,}")
    print(f"  Keywords analysed: {len(df_merged):,}")
    print(f"  Already in title: {in_title:,}")
    print(f"  Missing from title: {not_in_title:,}")

    # Save results
    if args.excel and OPENPYXL_AVAILABLE:
        output_path = (
            args.output.replace(".csv", ".xlsx")
            if args.output.endswith(".csv")
            else args.output + ".xlsx"
        )
        save_excel_highlighted(df_merged, output_path, sheet_name="Keyword Gap")
    else:
        df_merged.to_csv(args.output, index=False, encoding="utf-8-sig")
        print(f"\nCSV saved to: {args.output}")

    # Show top gaps
    df_gaps = df_merged[~df_merged["in_title"]].head(10)
    if len(df_gaps) > 0:
        print("\nTop keyword gaps (high impressions, missing from title):")
        for _, row in df_gaps.iterrows():
            print(
                f"  [{row['impressions']:,.0f} impr] "
                f"'{row['query']}' - {row['page'][:60]}"
            )


# ---------------------------------------------------------------------------
# Mode 2: Title Segment Analysis
# ---------------------------------------------------------------------------


def run_segment_analysis(args):
    """Segment mode: split titles, cross-reference against GSC data."""
    df_titles, df_queries = prepare_inputs(args)

    df_queries["kw_source"] = "gsc"

    # Parse brand terms for segment filtering
    brand_terms = []
    if args.brand:
        brand_terms = [b.strip().lower() for b in args.brand.split(",") if b.strip()]

    # -- Split titles into segments --
    df_segments = df_titles.copy()
    split_cols = (
        df_segments["title"]
        .str.split(args.delimiter, expand=True)
        .add_prefix("title_")
    )
    df_segments = df_segments.join(split_cols)

    segment_cols = [c for c in df_segments.columns if c.startswith("title_")]
    df_segments["query"] = df_segments[segment_cols].values.tolist()
    df_segments = df_segments[["page", "title", "query"]].explode("query")

    # Clean segment queries
    df_segments["query"] = df_segments["query"].astype(str).str.strip().str.lower()
    df_segments["query"] = df_segments["query"].str.split().str.join(" ")

    # Remove full-title and URL rows
    df_segments["title_lower"] = df_segments["title"].str.lower()
    df_segments = df_segments[df_segments["query"] != df_segments["title_lower"]]
    df_segments = df_segments[df_segments["query"] != df_segments["page"]]
    df_segments = df_segments.drop(columns=["title_lower"])

    # Remove empty/none values
    df_segments = df_segments[df_segments["query"].notna()]
    df_segments = df_segments[~df_segments["query"].isin(["", "none", "nan"])]

    # Filter brand terms from segments
    for term in brand_terms:
        df_segments = df_segments[~df_segments["query"].str.contains(term, na=False)]

    df_segments["kw_source"] = "page_title"

    print(f"\n  Extracted {len(df_segments):,} title segments from {df_titles['page'].nunique():,} pages")

    # -- Merge segment keywords with GSC data --
    df_seg_merged = pd.merge(
        df_segments,
        df_queries[["query", "page", "clicks", "impressions"]],
        on=["query", "page"],
        how="left",
    )
    cols_order = ["page", "query", "kw_source", "clicks", "impressions"]
    df_seg_merged = df_seg_merged.reindex(columns=cols_order)

    # -- Top GSC keywords --
    df_gsc_top = df_queries.copy()
    df_gsc_top = df_gsc_top.sort_values("clicks", ascending=False)
    df_gsc_top = df_gsc_top[df_gsc_top["clicks"] > 0]
    df_gsc_top = df_gsc_top.groupby("page").head(args.max_keywords)

    # -- Combine --
    df_combined = pd.concat(
        [df_seg_merged, df_gsc_top[cols_order]], ignore_index=True
    )
    df_combined.fillna({"clicks": 0, "impressions": 0}, inplace=True)
    df_combined.drop_duplicates(subset=["query", "page"], keep="first", inplace=True)
    df_combined.sort_values(["page", "clicks"], ascending=[True, False], inplace=True)

    # Attach title
    df_combined = pd.merge(
        df_combined, df_titles[["page", "title"]], on="page", how="left"
    )
    df_combined.drop_duplicates(subset=["page", "query"], keep="first", inplace=True)
    df_combined = df_combined[df_combined["title"].notna()]

    # Aggregates
    df_combined["total_clicks"] = df_combined.groupby("page")["clicks"].transform("sum")
    df_combined["total_impressions"] = df_combined.groupby("page")[
        "impressions"
    ].transform("sum")
    df_combined.sort_values(
        by=["total_impressions", "page"], ascending=[False, True], inplace=True
    )

    # Match status
    def check_match(row):
        query = str(row["query"]).strip().lower()
        title = str(row["title"]).strip().lower()
        segments = [s.strip().lower() for s in title.split(args.delimiter)]
        return any(query in seg for seg in segments)

    df_combined["in_title"] = df_combined.apply(check_match, axis=1)

    # Summary
    seg_count = (df_combined["kw_source"] == "page_title").sum()
    no_vol = (
        (df_combined["kw_source"] == "page_title") & (df_combined["impressions"] == 0)
    ).sum()
    gsc_missing = (
        (df_combined["kw_source"] == "gsc") & (~df_combined["in_title"])
    ).sum()

    print(f"\nResults:")
    print(f"  Pages analysed: {df_combined['page'].nunique():,}")
    print(f"  Total keywords: {len(df_combined):,}")
    print(f"  Title segments: {seg_count:,}")
    print(f"  Segments with no search volume: {no_vol:,}")
    print(f"  GSC keywords not in title: {gsc_missing:,}")

    # Save
    if args.excel and OPENPYXL_AVAILABLE:
        output_path = (
            args.output.replace(".csv", ".xlsx")
            if args.output.endswith(".csv")
            else args.output + ".xlsx"
        )
        save_excel_highlighted(df_combined, output_path, sheet_name="Segment Analysis")
    else:
        df_combined.to_csv(args.output, index=False, encoding="utf-8-sig")
        print(f"\nCSV saved to: {args.output}")

    # Show top issues
    df_no_vol = df_combined[
        (df_combined["kw_source"] == "page_title") & (df_combined["impressions"] == 0)
    ].head(5)
    if len(df_no_vol) > 0:
        print("\nTitle segments with no search volume (sample):")
        for _, row in df_no_vol.iterrows():
            print(f"  '{row['query']}' on {row['page'][:60]}")

    df_missing = df_combined[
        (df_combined["kw_source"] == "gsc") & (~df_combined["in_title"])
    ].head(5)
    if len(df_missing) > 0:
        print("\nTop GSC keywords missing from title (sample):")
        for _, row in df_missing.iterrows():
            print(
                f"  [{row['clicks']:,.0f} clicks] "
                f"'{row['query']}' - {row['page'][:60]}"
            )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Title Keyword Gap Finder v2 - Compare GSC keywords vs page titles"
    )
    subparsers = parser.add_subparsers(dest="mode", help="Analysis mode")

    # Shared arguments added to both subparsers
    def add_common_args(sub):
        sub.add_argument(
            "--crawl", required=True, help="Path to Screaming Frog crawl CSV"
        )
        sub.add_argument("--gsc", required=True, help="Path to GSC query data CSV")
        sub.add_argument(
            "--output",
            default="title_keyword_gaps.csv",
            help="Output file path (default: title_keyword_gaps.csv)",
        )
        sub.add_argument(
            "--delimiter", default="|", help="Title delimiter for splitting (default: |)"
        )
        sub.add_argument(
            "--brand",
            default="",
            help="Brand terms to exclude (comma-separated)",
        )
        sub.add_argument(
            "--url-filter",
            default="",
            help="Only analyse URLs containing this text",
        )
        sub.add_argument(
            "--max-keywords",
            type=int,
            default=10,
            help="Max keywords per page (default: 10)",
        )
        sub.add_argument(
            "--min-impressions",
            type=int,
            default=0,
            help="Minimum impressions threshold (default: 0)",
        )
        sub.add_argument(
            "--excel", action="store_true", help="Export as Excel with highlighting"
        )

    # Mode 1: keyword-gap
    parser_gap = subparsers.add_parser(
        "keyword-gap",
        help="Find GSC keywords driving traffic but missing from page titles",
    )
    add_common_args(parser_gap)

    # Mode 2: segment
    parser_seg = subparsers.add_parser(
        "segment",
        help="Split titles by delimiter, compare segments against GSC data",
    )
    add_common_args(parser_seg)

    args = parser.parse_args()

    if args.mode is None:
        parser.print_help()
        sys.exit(1)

    if args.mode == "keyword-gap":
        run_keyword_gap(args)
    elif args.mode == "segment":
        run_segment_analysis(args)


if __name__ == "__main__":
    main()
